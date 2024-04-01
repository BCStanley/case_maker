from __future__ import annotations

import os.path
import re
from typing import Optional
from openpyxl.cell import Cell
import sqlite3
from sqlite3 import Error
from itertools import chain

from openpyxl.reader.excel import load_workbook

from sql_structure import Table
from sql_structure import DatabaseStructure
from sql_structure import SelectionQuery


class Case:
    def __init__(
            self,
            name: str,
            court: str,
            year: Optional[int] = None,
            related_cases: Optional[list[str]] = None,
            nom_cite: Optional[str] = None,
            er_cite: Optional[str] = None,
            comment: Optional[str] = None,
            link: Optional[str] = None,
            area_tags: Optional[list[str]] = None,
            subject_tags: Optional[list[str]] = None,
            cite_in_tags: Optional[dict] = None,
            authors: Optional[list[str]] = None,
            special_terms: Optional[list[str]] = None,
            id: Optional[int] = None
    ):
        self.name = name
        self.year = year
        self.link = link
        self.court = court
        self.er_cite = er_cite
        self.nom_cite = nom_cite
        self.comment = comment
        self.area_tags = area_tags
        self.subject_tags = subject_tags
        self.cite_in_tags = cite_in_tags
        self.authors = authors
        self.special_terms = special_terms
        self.related_cases = related_cases
        self.id = id

    def display(self):
        print(f"""---
        {self.display_name}
        area_tags: {self.display_area_tags}
        subject_tags: {self.display_subject_tags}
        authors: {self.display_author}
        special_terms: {self.display_special_terms}
        comment: {self.comment}
        cited_in: {self.display_cite_ins}
        """)

    @property
    def display_name(self) -> str:
        return f'{self.name} ({self.year}) {self.nom_cite}, {self.er_cite}'

    @property
    def display_area_tags(self) -> str:
        return ', '.join(self.area_tags)

    @property
    def display_subject_tags(self) -> str:
        return ', '.join(self.subject_tags)

    @property
    def display_author(self) -> str:
        return ', '.join(self.authors)

    @property
    def display_special_terms(self) -> str:
        return ', '.join(self.special_terms)

    @property
    def display_cite_ins(self) -> str:
        return '; '.join([f'{person} ({comment})' for person, comment in self.cite_in_tags.items()])

    @property
    def as_dict(self) -> dict:
        return_dict = {"name": f"\"{self.name}\"",
                       "year": f"{str(self.year)}",
                       "nom_cite": f"\"{self.nom_cite}\"",
                       "er_cite": f"\"{self.er_cite}\"",
                       "court": f"\"{self.court}\"",
                       "link": f"\"{self.link}\"",
                       "comment": f"\"{self.comment}\""}
        return_dict = {k: v for k, v in return_dict.items() if v}
        return return_dict

    @staticmethod
    def from_excel(row: tuple) -> Case:
        def get_str_value(cell: Cell) -> str | None:
            return cell.value.strip() if cell.value else None

        def get_year(cell: Cell) -> int | None:
            if isinstance((date := cell.value), int):
                return date
            else:
                raise Exception(f'Error! Date at row {cell.row} is of type {type(date)}!')

        def get_underscore_list(cell: Cell) -> list[str] | None:
            if input_str := cell.value:
                return [item.replace('_', ' ') for item in input_str.split()]

        def get_cite_ins(cell) -> dict | None:
            def get_items(input_str: str) -> tuple[str, str]:
                pattern = re.compile(r'^([^[]+)\[([^]]+)\]$')
                match = pattern.match(input_str)
                if match:
                    return match.group(1), match.group(2)
                else:
                    print('No match found.')

            if cell.value:
                input_strings = cell.value.strip().split('; ')
                out_dict = dict()
                for input_str in input_strings:
                    person, comment = get_items(input_str)
                    out_dict.update({person: comment})
                return out_dict

        def get_comma_list(cell: Cell) -> list[str] | None:
            if input_str := cell.value:
                return input_str.strip().split(', ')

        return Case(
            name=get_str_value(row[0]),
            year=get_year(row[1]),
            nom_cite=get_str_value(row[2]),
            er_cite=get_str_value(row[3]),
            court=get_str_value(row[4]),
            subject_tags=get_underscore_list(row[5]),
            authors=get_underscore_list(row[6]),
            related_cases=get_comma_list(row[7]),
            cite_in_tags=get_cite_ins(row[8]),
            comment=get_str_value(row[9]),
            area_tags=get_underscore_list(row[10]),
            link=get_str_value(row[11]),
            special_terms=get_comma_list(row[12])
        )

    @staticmethod
    def alter_case(row: tuple) -> Case():
        def get_str_value(cell: Cell) -> str | None:
            return cell.value.strip() if cell.value else None

        def get_year(cell: Cell) -> int | None:
            if isinstance((date := cell.value), int):
                return date
            else:
                return

        def get_underscore_list(cell: Cell) -> list[str] | None:
            if input_str := cell.value:
                return [item.replace('_', ' ') for item in input_str.split()]
            else:
                return

        def get_cite_ins(cell) -> dict | None:
            def get_items(input_str: str) -> tuple[str, str]:
                pattern = re.compile(r'^([^[]+)\[([^]]+)\]$')
                match = pattern.match(input_str)
                if match:
                    return match.group(1), match.group(2)
                else:
                    print('No match found.')

            if cell.value:
                input_strings = cell.value.strip().split('; ')
                out_dict = dict()
                for input_str in input_strings:
                    person, comment = get_items(input_str)
                    out_dict.update({person: comment})
                return out_dict

        def get_comma_list(cell: Cell) -> list[str] | None:
            if input_str := cell.value:
                return input_str.strip().split(', ')

        return Case(
            name=get_str_value(row[0]),
            year=get_year(row[1]),
            nom_cite=get_str_value(row[2]),
            er_cite=get_str_value(row[3]),
            court=get_str_value(row[4]),
            subject_tags=get_underscore_list(row[5]),
            authors=get_underscore_list(row[6]),
            related_cases=get_comma_list(row[7]),
            cite_in_tags=get_cite_ins(row[8]),
            comment=get_str_value(row[9]),
            area_tags=get_underscore_list(row[10]),
            link=get_str_value(row[11]),
            special_terms=get_comma_list(row[12])
        )


class Casebook:

    def __init__(
            self,
            sql_path: str,
            cases: list[Case] = [],
    ):
        self.cases = cases
        self.sql_path = sql_path
        try:
            self.sql_connection: sqlite3.Connection = sqlite3.connect(self.sql_path)
            print(f"Connection to {self.sql_path} successful.")
        except Error as e:
            print(f"The error {e} occurred")
            raise ValueError(f"The file {self.sql_path} does not exist, or could not be connected to.")
        self.database = DatabaseStructure()

    def execute(self, query: str):
        cursor = self.sql_connection.cursor()
        try:
            cursor.execute("PRAGMA foreign_keys = ON")
            cursor.execute(query)
            self.sql_connection.commit()
            print("Query executed successfully")
        except Error as e:
            print(f"The error '{e}' occurred")

    @property
    def display_casebook_info(self) -> str:
        """Gives a basic display of the key information about the Casebook object.
        :return: str, number of cases [number], SQL database [path].
        """
        return f"This casebook has {len(self.cases)} cases, with an SQL file at {self.sql_path}."
        pass

    def find_case_by_id(self, given_id):
        for case_object in self.cases:
            if case_object.id == given_id:
                return case_object
            else:
                pass

    @staticmethod
    def make_new_database(temp_connection: sqlite3.Connection):
        """ A static method for producing a new database, which is called as part of the new_casebook_from_xl and
        new_casebook_from_sql functions.
        :param temp_connection: this must be a sqlite3.Connection object, which will already have been established.
        :return: None, but it inserts the relevant information.
        """
        all_tables = DatabaseStructure()
        cursor = temp_connection.cursor()
        for table in all_tables:
            try:
                cursor.execute("PRAGMA foreign_keys = ON")
                cursor.execute(table.creation_query)
                temp_connection.commit()
                print(f"Table {table.title} created successfully")
            except Error as e:
                print(f"Failed to create table {table.title}. The error '{e}' occurred")

    @staticmethod
    def enter_cases(temp_connection: sqlite3.Connection, list_of_cases: list[Case]) -> int | list[Case]:
        """
        Takes a connection object and a list of cases (taken from elsewhere.) These are entered into the database,
        and a new list is returned, which includes the "id" field, taken from where these are in the database.
        :param temp_connection: a sqlite3.Connection object.
        :param list_of_cases: a list of Case() objects, missing their id fields (or with what will be incorrect ids).
        :return: a new list[Case], where each case has been updated.
        """

        def enter_crossref_fields(case_object: Case, case_object_item: Case.something,
                                  info_table: Table, info_crossref_table: Table) -> None:
            """
            This is a bit of a monster of a function. What it does is deal with any of the fields in a case which involve
            crossref tables of some sort, checking to see if a value is already there and then entering it if it is.
            The relevant crossref data is there. NOTE: a different function is called for those that use comments.
            :param case_object: a Case() object.
            :param case_object_item: an object within the Case() object that one is looking to enter.
            :param info_table: The corresponding table in the DatabaseStructure()
            :param info_crossref_table: The corresponding crossref table in the DatabaseStructure()
            :return: Nothing is returned.
            """
            cursor = temp_connection.cursor()
            if case_object_item is None:
                return
            if info_table.title == "cited_in":
                return
            for name in case_object_item:  # First, test if the relevant item has already been inserted.
                cursor.execute(
                    f"SELECT {list(info_table.fields)[0]} from {info_table.title} WHERE {list(info_table.fields)[1]} = \"{name}\";")
                object_id = cursor.fetchone()
                if object_id:  # If it has, skip and just add the crossref information.
                    object_id = object_id[0]
                    dict_for_creation = {list(info_crossref_table.fields)[0]: str(object_id),
                                         list(info_crossref_table.fields)[1]: str(case.id)}
                    cursor.execute(info_crossref_table.insert_query(dict_for_creation))
                else:  # Otherwise, add the entry to the main table.
                    dict_for_creation = {list(info_table.fields)[1]: f"\"{name}\""}
                    cursor.execute(
                        info_table.insert_query(dict_for_creation))  # Then, add the relevant crossref entries.
                    cursor.execute(
                        f"SELECT {list(info_table.fields)[0]} from {info_table.title} WHERE {list(info_table.fields)[1]} = \"{name}\";")
                    object_id = cursor.fetchone()[0]
                    dict_for_creation = {list(info_crossref_table.fields)[0]: str(object_id),
                                         list(info_crossref_table.fields)[1]: str(case_object.id)}
                    cursor.execute(info_crossref_table.insert_query(dict_for_creation))

        def enter_crossref_fields_with_comments(case_object: Case, case_object_item: Case.something,
                                                info_table: Table, info_crossref_table: Table) -> None:
            """
            This is a modification of the enter_crossref_fields function (see the documentation for that) for those
            entry types based on a dict, meaning a different procedure is needed (mainly for the "cited_in" fields).
            :param case_object: a Case() object.
            :param case_object_item: an object within the Case() object that one is looking to enter.
            :param info_table: The corresponding table in the DatabaseStructure()
            :param info_crossref_table: The corresponding crossref table in the DatabaseStructure()
            :return: Nothing is returned.
            """
            cursor = temp_connection.cursor()
            if case_object_item is None:
                return
            for item_key in case_object_item:  # First we need to test if the "item_key" is there.
                cursor.execute(
                    f"SELECT {list(info_table.fields)[0]} from {info_table.title} WHERE {list(info_table.fields)[1]} = \"{item_key}\";")
                item_id = cursor.fetchone()
                if item_id:  # If so, we skip and add the crossref information.
                    item_id = item_id[0]
                    dict_for_creation = {list(info_crossref_table.fields)[0]: str(item_id),
                                         list(info_crossref_table.fields)[1]: str(case_object.id),
                                         list(info_crossref_table.fields)[2]: f"\"{case_object_item[item_key]}\""}
                    cursor.execute(info_crossref_table.insert_query(dict_for_creation))
                else:  # Otherwise, we need to add the relevant author.
                    dict_for_creation = {list(info_table.fields)[1]: f"\"{item_key}\""}
                    cursor.execute(info_table.insert_query(dict_for_creation))  # And then add the relevant crossref.
                    cursor.execute(
                        f"SELECT {list(info_table.fields)[0]} from {info_table.title} WHERE {list(info_table.fields)[1]} = \"{item_key}\";")
                    item_id = cursor.fetchone()[0]
                    dict_for_creation = {list(info_crossref_table.fields)[0]: str(item_id),
                                         list(info_crossref_table.fields)[1]: str(case_object.id),
                                         list(info_crossref_table.fields)[2]: f"\"{case_object_item[item_key]}\""}
                    cursor.execute(info_crossref_table.insert_query(dict_for_creation))

        all_tables = DatabaseStructure()
        cursor = temp_connection.cursor()
        list_of_cases_with_ids = []
        for case in list_of_cases:
            try:  # First, add the relevant case to the database.
                cursor.execute("PRAGMA foreign_keys = ON")
                cursor.execute(all_tables.cases_table.insert_query(case.as_dict))
                print(f"case {case.display_name} entered.")
            except Error as e:
                print(f"Failed to create case {case.display}. The error '{e}' occurred")
            try:  # Find the "id" as an int, and then add this to the list to be returned.
                cursor.execute(f"SELECT id from cases WHERE name=\"{case.name}\";")
                case_id: int = cursor.fetchone()[0]
                case.id = case_id
                list_of_cases_with_ids.append(case)
                enter_crossref_fields(case, case.subject_tags, all_tables.subject_table,
                                      all_tables.subject_crossref_table)
                enter_crossref_fields(case, case.authors, all_tables.authors_table, all_tables.authors_crossref_table)
                enter_crossref_fields(case, case.area_tags, all_tables.area_table, all_tables.area_crossref_table)
                enter_crossref_fields(case, case.special_terms, all_tables.special_terms_table,
                                      all_tables.special_terms_crossref_table)
                enter_crossref_fields_with_comments(case, case.cite_in_tags, all_tables.cite_in_table,
                                                    all_tables.cite_in_crossref_table)
            except Error as e:
                print(f"Could not find case {case.name} in database. The error '{e}' occurred.")
        temp_connection.commit()
        return list_of_cases_with_ids  # return the list of cases, with the new ids.

    @staticmethod
    def new_casebook_from_xl(sql_path: str, xl_path: str) -> Casebook | None:
        """ Create a new Casebook with a .db file of the path, sql_path.
        Enter new data into that casebook (and likewise the .db file) from the xl_path. 
        :param sql_path: a str object, a path to a blank or non-existent .db file. 
        :param xl_path: a str object, a path to a .xlsx file containing the relevant data. 
        :return: a Casebook object, with cases generated from .xlsx file (using Case.from_excel() for each row), and a
        sql connection object generated from sql_path. 
        """

        def read_cases_from_new(xl: str = xl_path) -> list[Case]:
            """
            This reads the relevant xl_file, and produces the relevant list[Case] needed to __init__ a new Casebook.
            :param xl: a str, which is a path to a .xlsx file. The function will only be engaged when that file
            is known to exist.
            :return: a list object, containing Case objects.
            """
            sheet = load_workbook(xl_path)["New_Cases"]
            new_list = []
            for row in sheet.iter_rows(min_row=2, max_col=13):
                if row[0].value == "DONE":
                    break
                else:
                    new_list.append(Case.from_excel(row=row))
            return new_list

        if os.path.isfile(sql_path):  # Establish whether there is already a .db file at the sql_path.
            print(f"The file {sql_path} already exists. Cannot make new database.")  # If so, raise an error.
            return None  # And return without doing anything.
        else:  # If there is no .db file, proceed to make one.
            try:
                new_sql_connection: sqlite3.Connection = sqlite3.connect(sql_path)
                print(f"Connection to {sql_path} successful after being made.")
            except Error as e:  # If there is an error in doing this, then raise the error.
                print(f"The error {e} occurred")
                return None  # And return without doing anything.
        if os.path.isfile(xl_path):  # Establish whether the .xlsx file exists
            Casebook.make_new_database(new_sql_connection)  # Then make the new database.
            new_list = read_cases_from_new(xl_path)  # Then make list of cases.
            new_list_with_ids = Casebook.enter_cases(new_sql_connection, new_list)
            print("function finished")
            return Casebook(sql_path, new_list_with_ids)
        else:
            print(f"The file {xl_path} does not exist.")
            return None

    @staticmethod
    def new_casebook_from_sql(sql_path: str) -> Casebook | None:
        """ Create a new Casebook with a .db file of the path, sql_path.
        The data is read from the sql file itself, which already contains the relevant data.
        :param sql_path: a str object, a path to a .db file containing relevant data.
        :return: a Casebook object, with cases generated from the .db file and a sql connection object (sql_path).
        """

        def get_data(case_id: int, table: Table, crossref_table: Table) -> list[str]:
            """
            A function, used within the new_casebook_from_sql method, which obtains from the database the relevant data
            in a form that can be entered into a Case() object.
            :param case_id: an int, which corresponds with a case id in the database.
            :param table: a Table() object, the Table where the information is sought from, e.g. subjects.
            :param crossref_table: a Table() object from the DatabaseStructure() object, corresponding to the table given.
            e.g., if subjects where given, subjects_crossref.
            :return: a list of str, in the form that is appropriate for a Case() object.
            """
            relevant_database_query = SelectionQuery(
                table,
                ["name"],
                {"id": ["in", SelectionQuery(crossref_table, [list(crossref_table.fields.keys())[0]],
                                             {list(crossref_table.fields.keys())[1]: ["=", case_id]}).sql_text]}
            )
            cursor.execute(relevant_database_query.full_sql_text)
            return list(chain.from_iterable(cursor.fetchall()))

        def get_data_with_comments(case_id: int, table: Table, crossref_table: Table) -> dict:
            """
            A function for obtaining data for the comments section of a Case() object. This is essentially a modified
            version of get_data() function.
            :param case_id: case_id: an int, which corresponds with a case id in the database
            :param table: a Table() object, the Table where the information is sought from, e.g. subjects.
            :param crossref_table: a Table() object from the DatabaseStructure() object, corresponding to the table given.
            :return: a dict(), in the format needed for the Case() object.
            """
            first_database_query = SelectionQuery(
                # A SelectionQuery() object which selects the ids of the subjects etc.
                crossref_table,
                [list(crossref_table.fields)[0], list(crossref_table.fields)[2]],
                {list(crossref_table.fields)[1]: ["=", case_id]}
            )
            cursor.execute(first_database_query.full_sql_text)
            return_dict = {}
            for result in cursor.fetchall():  # Iterate through all of those subject ids.
                comment = result[1]
                person_id = result[0]
                additional_database_query = SelectionQuery(  # And then obtain the names of those.
                    table,
                    ["name"],
                    {"id": ["=", person_id]}
                )
                cursor.execute(additional_database_query.full_sql_text)
                person_name = cursor.fetchone()[0]
                return_dict[person_name] = comment
            return return_dict

        if not os.path.isfile(sql_path):  # First check if the file exists.
            print(f"The database {sql_path} does not exist")
            return
        try:
            connection: sqlite3.Connection = sqlite3.connect(sql_path)
            print(f"Connection to {sql_path} successful.")
        except Error as e:
            print(f"The error {e} occurred")
        all_tables = DatabaseStructure()
        cursor = connection.cursor()
        get_all_cases = SelectionQuery(all_tables.cases_table, ["*"], {})  # Obtain the cases.
        cursor.execute(get_all_cases.full_sql_text)
        obtained_cases = cursor.fetchall()
        cases_list: list = []
        for case in obtained_cases:  # Then iterate through them.
            new_case = Case(  # And create the new Case() object.
                id=case[0],
                name=case[1],
                year=case[2],
                nom_cite=case[3],
                er_cite=case[4],
                court=case[5],
                link=case[6],
                comment=case[7],
                area_tags=get_data(case[0], all_tables.area_table, all_tables.area_crossref_table),
                subject_tags=get_data(case[0], all_tables.subject_table, all_tables.subject_crossref_table),
                authors=get_data(case[0], all_tables.authors_table, all_tables.authors_crossref_table),
                special_terms=get_data(case[0], all_tables.special_terms_table,
                                       all_tables.special_terms_crossref_table),
                cite_in_tags=get_data_with_comments(case[0], all_tables.cite_in_table,
                                                    all_tables.cite_in_crossref_table)
            )
            cases_list.append(new_case)
        return Casebook(sql_path, cases_list)  # Return a Casebook() object.

    def update_casebook_from_xl(self, xl_path: str) -> None:
        """ Updates the Casebook object with data given from a .xlsx file, at xl_path.
        :param xl_path: a str object, a path to a .xlsx file containing the relevant data.
        :return: self, the Casebook object is itself updated.
        """

        all_tables = DatabaseStructure()

        def check_values(found: str | int, existing: str | int, column: str, existing_id: int) -> str | None:
            """
            For checking and updating values in the case table without use of the crossref list.
            This means that only one alteration is needed, this always being to the cases_table.
            :param found: the value found in the alterations table, either an int or a str.
            :param existing: the value found in the existing case, either an int or a str.
            :param column: the column that is going to be altered in the table.
            :param: existing_id, the id of the case being altered.
            :return: a str, which is the SQL text to execute to alter the case data or None if no changes are needed.
            """
            if not found:
                return  # If no value is found, return without doing anything.
            else:  # If the value found is the same as that already given, then return without doing anything.
                if found == existing:
                    return
                else:  # If something is found, and it is different, return the relevant update text.
                    return all_tables.cases_table.update_query({column: found}, ("id", existing_id))

            def check_values_with_crossref(info_to_add: list, crossref_table: Table, existing_id: int) -> str | None:
                pass

        sheet = load_workbook(xl_path)["Alterations"]
        for row in sheet.iter_rows(min_row=2, max_col=14):  # Iterate through the Alterations part of the table.
            if row[0].value == "DONE":  # Break if it has reached the end of the entries.
                break
            else:
                alter_name: str = row[0].value  # Find the name given for the case to be altered.
                is_case_query = SelectionQuery(all_tables.cases_table, ["id"], {"name": ["=", alter_name]})
                cursor = self.sql_connection.cursor()
                cursor.execute(
                    is_case_query.full_sql_text)  # Try and find that name in the cases_table, looking for id.
                found_id = cursor.fetchone()
                if found_id:  # If the id is found, then begin to make the alterations.
                    # Make a Case object, which contains the relevant alterations.
                    test_case: Case = Case.alter_case(row=row[1:14])
                    # Pull up the Case object which corresponds with the id that has been found.
                    found_case = self.find_case_by_id(found_id[0])
                    # Now, we need to go through each of the attributes of the Case object and see if alterations are needed.
                    # name
                    print(check_values(test_case.name, found_case.name, "name", found_id[0]))
                    # year
                    print(check_values(test_case.year, found_case.year, "year", found_id[0]))
                    # link
                    print(check_values(test_case.link, found_case.link, "link", found_id[0]))
                    # court
                    print(check_values(test_case.court, found_case.court, "court", found_id[0]))
                    # er_cite
                    print(check_values(test_case.er_cite, found_case.er_cite, "er_cite", found_id[0]))
                    # nom_cite
                    print(check_values(test_case.nom_cite, found_case.nom_cite, "nom_cite", found_id[0]))
                    # comment
                    print(check_values(test_case.comment, found_case.comment, "comment", found_id[0]))
                    # These next few ones are a bit more complex, because two sets of tables need to be updated.
                    # area_tags
                    # subject_tags
                    # authors
                    # special_terms
                    # comment
                else:  # Otherwise, do nothing.
                    pass
