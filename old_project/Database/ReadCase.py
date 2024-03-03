from openpyxl import Workbook
from openpyxl import load_workbook
import sqlite3
from sqlite3 import Error
import sys


list = "Input_Sheet.xlsx"

# This function returns a tuple, which is a list of lists of the lines of the spreadsheet given as "spreadsheet_path"
def obtain_list_from_spreadsheet(spreadsheet_path):
    workbook = load_workbook(spreadsheet_path)
    sheet = workbook["New_Cases"]
    has_break_value = False # A boolean variable that is used to find the "break value"
    # This simply looks to find the relevant break value.
    for row in sheet.iter_rows(min_row=1, max_row=1000, min_col=1, max_col=1):
        # The effect of the above will mean that the theoretical maximum number is 10000.
        for i in row:
            current_value = str(i.value)
            if current_value == "DONE":
                break_value = i
                has_break_value = True
            else:
                pass
    if has_break_value == False:
        max_value = 50 # This means that if no "DONE" string is found, the set break value will be 50.
    else:
        max_value = break_value.row -1 # The "max_value" is set to the "break_value".
    # Having obtained the max value, the next step is fairly straightforward. All that needs to be done is an interation through the table to obain the relevant strings.
    tuple = [] # This tuple will be the full item that is returned.
    for row in sheet.iter_rows(min_row=2, max_row=max_value, min_col=1, max_col=13):
        row_list = [] # This variable resets itself every time a pass through is done. The point is to create a list which is then added to the tuple after each row.
        for i in row:
            row_list.append(str(i.value))
        tuple.append(row_list)
    return tuple

# A very simple function which establishes a connection to the relevant database.
def create_connection(path):
    connection = None
    try:
        connection = sqlite3.connect(path)
        print("Connection to SQLite DB successful")
    except Error as e:
        print(f"The error '{e}' occurred")
    return connection


# This very simple function executes the relevant query on the databse.
def execute_query(connection, query):
    cursor = connection.cursor()
    try:
        cursor.execute("PRAGMA foreign_keys = ON")
        cursor.execute(query)
        connection.commit()
        print("Query executed successfully")
    except Error as e:
        print(f"The error '{e}' occurred")

def make_new_table(database_path, table_data):
    current_connection = create_connection(database_path)
    execute_query(current_connection, table_data)


def sort_out_raw_entry(list, mode):
    if mode == "cases":
        output = []
        for line in list:
            del line[12]
            del line[5:8]
            del line[5]
            del line[6]
            line[0] = line[0].strip()
            output.append(line)
    elif mode == "subjects":
        output = dict()
        for line in list:
            name = line[0]
            name = name.strip()
            subjects = line[5]
            output[name] = subjects
    elif mode == "citing":
        output = dict()
        for line in list:
            name = line[0]
            name = name.strip()
            citing = line[6]
            output[name] = citing
    elif mode == "legal_area":
        output = dict()
        for line in list:
            name = line[0]
            name = name.strip()
            legal_area = line[10]
            output[name] = legal_area
    elif mode == "comment":
        output = dict()
        for line in list:
            name = line[0]
            name = name.strip()
            comment = line[9]
            output[name] = comment
    elif mode == "special_terms":
        output = dict()
        for line in list:
            name = line[0]
            name = name.strip()
            terms = line[12]
            output[name] = terms
    elif mode == "cited in":
        output = dict()
        for line in list:
            name = line[0]
            name = name.strip()
            cited_in = line[8]
            output[name] = cited_in
    elif mode == "cases_cited":
        output = dict()
        for line in list:
            name = line[0]
            name = name.strip()
            cited_in = line[7]
            output[name] = cited_in
    return output

# This function reads a list, inputed from obtain_list_from_spreadsheet() function, and creates a block of text suitable as executing into an SQL String.
def convert_raw_list_to_SQL_Q(value):
    total_lines = len(value)
    entry_block = ""
    line_counter = 0
    for line in value:
        line_list = ""
        n = 0
        for item in line:
            item = item.strip()
            if n != 6:
                line_list += "\'" + item + "\'" + ", "
            else:
                line_list += "\'" + item + "\'"
            n = n + 1
        line_list = "\t(" + line_list + ")"
        if line_counter < total_lines-1:
            entry_block = entry_block + line_list + "," + "\n"
        else:
            entry_block = entry_block + line_list + ";"
        line_counter = line_counter + 1
    return entry_block


def insert_cases_from_xl(spreadsheet_path, connection):
    # The first thing that is done is that a tuple is made by reading the relevant spreadsheet.
    raw_data_tuple = obtain_list_from_spreadsheet(spreadsheet_path)
    # Then clean this up to the relevant data structure.
    cleaned_up = sort_out_raw_entry(raw_data_tuple, "cases")
    raw_data_tuple = obtain_list_from_spreadsheet(spreadsheet_path)
    comments = sort_out_raw_entry(raw_data_tuple, "comment")
    # Then a connection is established
    established_connection = create_connection(connection)
    # Then the relevant block of entry material is created.
    query_text = convert_raw_list_to_SQL_Q(cleaned_up)
    creation_text = """
    INSERT INTO
        cases (name, year, nom_cite, er_cite, court, comment, link)
    VALUES
    """
    creation_text = creation_text + query_text
    print(creation_text)
    execute_query(established_connection, creation_text)

# This function simply obtains all of the cases as a list, and sorts them by date.
def obtain_all_cases(data_path):
    temp_connection = create_connection(data_path)
    cur = temp_connection.cursor()
    cur.execute("SELECT * FROM cases ORDER BY year;")
    rows = cur.fetchall()
    return rows


create_case_table = """
CREATE TABLE IF NOT EXISTS cases (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  name TEXT NOT NULL UNIQUE,
  year INTEGER,
  nom_cite TEXT,
  er_cite TEXT,
  court TEXT,
  link TEXT,
  comment TEXT
);
"""

create_subject_table = """
CREATE TABLE IF NOT EXISTS subjects (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  name TEXT NOT NULL UNIQUE
);
"""

create_subject_crosref_table = """
CREATE TABLE IF NOT EXISTS subjectsCrossref (
  subjectsId INTEGER REFERENCES subjects (id),
  caseID INTEGER REFERENCES cases(id),
  UNIQUE (subjectsId, caseID)
);
"""

create_authors_table = """
CREATE TABLE IF NOT EXISTS authors (
  authorID INTEGER PRIMARY KEY AUTOINCREMENT,
  name TEXT NOT NULL UNIQUE
);
"""

create_authors_crossref_table = """
CREATE TABLE IF NOT EXISTS authorsCrossref (
  authorID INTEGER REFERENCES authors (authorID),
  caseID INTEGER REFERENCES cases(id),
  comment TEXT
);
"""

create_citing_table = """
CREATE TABLE IF NOT EXISTS citing (
  author_citing_id INTEGER PRIMARY KEY AUTOINCREMENT,
  name TEXT NOT NULL UNIQUE
);
"""

create_citing_table_crossref = """
CREATE TABLE IF NOT EXISTS citingCrossref (
  caseID INTEGER REFERENCES cases(id),
  author_cited INTEGER REFERENCES citing(author_citing_id)
);
"""

create_legal_area_table = """
CREATE TABLE IF NOT EXISTS legalArea (
  legal_area_id INTEGER PRIMARY KEY AUTOINCREMENT,
  name TEXT NOT NULL UNIQUE
);
"""

create_legal_area_table_crossref = """
CREATE TABLE IF NOT EXISTS legalAreaCrossref (
  caseID INTEGER REFERENCES cases(id),
  leal_area_id INTEGER REFERENCES legalArea(legal_area_id)
);
"""

create_terms_table = """
CREATE TABLE IF NOT EXISTS terms (
  term_id INTEGER PRIMARY KEY AUTOINCREMENT,
  name TEXT NOT NULL UNIQUE
);
"""

create_terms_table_crossref = """
CREATE TABLE IF NOT EXISTS termsCrossref (
  caseID INTEGER REFERENCES cases(id),
  term_id INTEGER REFERENCES terms(term_id)
);
"""

def insert_subjects_crossref(data_path, subject, case):
    subjectID = ""
    # The first step is to take the "case" given and find that case in the "cases" table in the database, and return the "id"
    temp_connection = create_connection(data_path)
    cur = temp_connection.cursor()
    cur.execute("SELECT id FROM cases WHERE name=\"" + case + "\";")
    caseID_list = cur.fetchone()
    if caseID_list == None:
        print("Error from insert_subjects_crossref: case name \'" + case + "\' not found. Skipping.")
    else:
        caseID = caseID_list[0]
        cur.execute("SELECT id FROM subjects WHERE name=\"" + subject + "\";")
        subjectID_list = cur.fetchone()
        # If the subject line is not found in the subjects table, it will be added.
        if subjectID_list == None:
            subject_line_creation_text = "INSERT INTO subjects (name) VALUES (\"" + subject + "\");"
            execute_query(temp_connection, subject_line_creation_text)
        else:
            subjectID = str(subjectID_list[0])
        subject_crossref_creation_text = "INSERT INTO subjectsCrossref (subjectsId, caseID) VALUES (" + subjectID + ", " + str(caseID) + ");"
        execute_query(temp_connection, subject_crossref_creation_text)

# This function should take the xl_path, and return a dictionary consiting of {cases : [list, of, tags]}
def read_subject_tags_from_xl(xl_path):
    raw_list = obtain_list_from_spreadsheet(xl_path)
    case_list = sort_out_raw_entry(raw_list, "subjects")
    for case in case_list:
        case_list[case].strip()
        subject_list = case_list[case].split(" ")
        subject_list = [x for x in subject_list if len(x)>3]
        case_list[case] = subject_list
    return case_list


# This function should take the xl_path and call the function read_subject_tags_from_xl to obtain the required dictionary. It then cycles through that dictionary, calling insert_subjects_crossref to add the relevant crossref tags.
def enter_subject_tags_from_xl(xl_path, data_path):
    # First call the read_subject_tags_from_xl function to obtain the correct dictionary.
    dict = read_subject_tags_from_xl(xl_path)
    for case in dict:
        subject_tag_list = dict[case]
        for subject_tag in subject_tag_list:
            insert_subjects_crossref(data_path, subject_tag, case)

def generate_subject_tag_dict(data_path):
    new_dict = dict()
    temp_connection = create_connection(data_path)
    cur = temp_connection.cursor()
    cur.execute("SELECT * FROM subjects;")
    subject_list = cur.fetchall()
    for row in subject_list:
        subject_id = row[0]
        list_of_caseIDs = []
        cur = temp_connection.cursor()
        cur.execute("SELECT caseID from subjectsCrossref where subjectsId=" + str(subject_id) + ";")
        sub_case_list = cur.fetchall()
        # What gets returned are somewhat messy tuples. I've converted them to simple integers
        cleaned_up_sub_case_list = []
        for member in sub_case_list:
            value = member[0]
            cleaned_up_sub_case_list.append(value)
        new_dict[subject_id] = cleaned_up_sub_case_list
    return new_dict


def read_author_cites_from_xl(xl_path):
    # First, we need to obtain a list of all of the "authors", which is produced as a dictionary {case_name: [[author, comment], [author, comment]]...}. Both keys are strings.
    raw_list = obtain_list_from_spreadsheet(xl_path)
    raw_authors_list = sort_out_raw_entry(raw_list, "cited in")
    cases_by_authors_dict = {}
    for case in raw_authors_list:
        case = case.strip()
        raw_authors_list[case].strip()
        # The individual entries are split at a semicolon and then a space
        split_author_list = raw_authors_list[case].split("; ")
        split_author_list = [x for x in split_author_list if len(x)>3]
        if split_author_list[0] == "None":
            cases_by_authors_dict[case] = None
        else:
            # Then we need to go through each author citation, and split this in two: the citation and the comment.
            total_authors_list = []
            for author in split_author_list:
                author.strip()
                author_and_comment = author.split("[")
                author_and_comment[1] = author_and_comment[1].rstrip()
                author_and_comment[1] = (author_and_comment[1])[:-1]
                total_authors_list.append(author_and_comment)
            cases_by_authors_dict[case] = total_authors_list
    return cases_by_authors_dict


def enter_author_comments_from_xl(xl_path, data_path):
    #step 1: obtain the dictionary from the xl.
    dictionary = read_author_cites_from_xl(xl_path)
    # step 2: cycle through each case in the dict
    for case in dictionary:
    # 2.1. for each case, obtain the caseID from the list of cases.
        temp_connection = create_connection(data_path)
        cur = temp_connection.cursor()
        case = case.strip()
        cur.execute("SELECT id FROM cases where name=\"" + case + "\";")
        caseID = (cur.fetchone())[0]
    # step 3: for each case, cycle through each author-comment made.
        case_author_info = dictionary[case]
        if not case_author_info:
            pass
        else:
            for author_comment in case_author_info:
                author = author_comment[0]
                comment = author_comment[1]
    # step 4: check the "author" entry. If it's not in the existing list of authors, add it.
    # 4.1. call up the "author table"
                cur = temp_connection.cursor()
                cur.execute("SELECT authorID FROM authors where name=\"" + author + "\";")
                author_ID = (cur.fetchone())
    # 4.2. check if that value empty, and if not add the author to the table.
                if author_ID == None:
                    author_line_creation_text = "INSERT INTO authors (name) VALUES (\"" + author + "\");"
                    execute_query(temp_connection, author_line_creation_text)
                    print("From enter_author_comments_from_xl: I do not have the author " + author + " so I have added it.")
                    author_ID = cur.execute("SELECT authorID FROM authors where name=\"" + author + "\";")
                    author_ID = (cur.fetchone())
                else:
                    pass
    # step 5: once this is sorted out, add the relevant cross-reference to the crossref table, including the caseID and author ID and comment.
                author_ID = author_ID[0]
                author_commment_creation_text = "INSERT INTO authorsCrossref (authorID, caseID, comment) VALUES (" + str(author_ID) + ", " + str(caseID) + ", \" " + comment + "\");"
                execute_query(temp_connection, author_commment_creation_text)


def generate_author_comment_dict(data_path):
    new_dict = dict()
    temp_connection = create_connection(data_path)
    cur = temp_connection.cursor()
    cur.execute("SELECT * FROM authors;")
    author_list = cur.fetchall()
    for row in author_list:
        author_id = row[0]
        cur = temp_connection.cursor()
        cur.execute("SELECT caseID, comment from authorsCrossref where authorID=" + str(author_id) + ";")
        sub_case_list = cur.fetchall()
        # What gets returned are somewhat messy tuples. I've converted them to simple integers
        cleaned_up_sub_case_list = []
        for member in sub_case_list:
            author = member[0]
            comment = member[1]
            list_to_append = (author, comment)
            cleaned_up_sub_case_list.append(list_to_append)
        new_dict[author_id] = cleaned_up_sub_case_list
    return new_dict


def read_citing_from_xl(xl_path):
    raw_list = obtain_list_from_spreadsheet(xl_path)
    case_list = sort_out_raw_entry(raw_list, "citing")
    for case in case_list:
        case_list[case].strip()
        citing_list = case_list[case].split(" ")
        citing_list = [x for x in citing_list if len(x)>3]
        case_list[case] = citing_list
    return case_list


def insert_citing_crossref(data_path, citing, case):
    subjectID = ""
    # The first step is to take the "case" given and find that case in the "cases" table in the database, and return the "id"
    temp_connection = create_connection(data_path)
    cur = temp_connection.cursor()
    cur.execute("SELECT id FROM cases WHERE name=\"" + case + "\";")
    caseID_list = cur.fetchone()
    if caseID_list == None:
        print("Error from insert_citing_crossref: case name \'" + case + "\' not found. Skipping.")
    else:
        caseID = caseID_list[0]
        cur.execute("SELECT author_citing_id FROM citing WHERE name=\"" + citing + "\";")
        citingID_list = cur.fetchone()
        # If the subject line is not found in the subjects table, it will be added.
        if citingID_list == None:
            # We need to make sure that we don't add a "none" entry to the table.
            if citing == "None":
                return
            else:
                citing_line_creation_text = "INSERT INTO citing (name) VALUES (\"" + citing + "\");"
                execute_query(temp_connection, citing_line_creation_text)
                cur.execute("SELECT author_citing_id FROM citing WHERE name=\"" + citing + "\";")
                citingID_list = cur.fetchone()
                citingID = str(citingID_list[0])
        else:
            citingID = str(citingID_list[0])
        citing_crossref_creation_text = "INSERT INTO citingCrossref (caseID, author_cited) VALUES (" + str(caseID) + ", " + citingID + ");"
        execute_query(temp_connection, citing_crossref_creation_text)


def enter_citing_from_xl(xl_path, data_path):
    dict = read_citing_from_xl(xl_path)
    for case in dict:
        citing_list = dict[case]
        for cite in citing_list:
            insert_citing_crossref(data_path, cite, case)


def generate_citing_dict(data_path):
    new_dict = dict()
    temp_connection = create_connection(data_path)
    cur = temp_connection.cursor()
    cur.execute("SELECT * FROM citing;")
    citing_list = cur.fetchall()
    for row in citing_list:
        citing_id = row[0]
        list_of_caseIDs = []
        cur = temp_connection.cursor()
        cur.execute("SELECT caseID from citingCrossref where author_cited=" + str(citing_id) + ";")
        sub_case_list = cur.fetchall()
        # What gets returned are somewhat messy tuples. I've converted them to simple integers
        cleaned_up_sub_case_list = []
        for member in sub_case_list:
            value = member[0]
            cleaned_up_sub_case_list.append(value)
        new_dict[citing_id] = cleaned_up_sub_case_list
    return new_dict

def read_legal_area_from_xl(xl_path):
    raw_list = obtain_list_from_spreadsheet(xl_path)
    case_list = sort_out_raw_entry(raw_list, "legal_area")
    for case in case_list:
        case_list[case].strip()
        area_list = case_list[case].split(" ")
        area_list = [x for x in area_list if len(x)>3]
        case_list[case] = area_list
    return case_list

def insert_legal_area_crossref(data_path, area, case):
    subjectID = ""
    # The first step is to take the "case" given and find that case in the "cases" table in the database, and return the "id"
    temp_connection = create_connection(data_path)
    cur = temp_connection.cursor()
    cur.execute("SELECT id FROM cases WHERE name=\"" + case + "\";")
    caseID_list = cur.fetchone()
    if caseID_list == None:
        print("Error from insert_legal_area_crossref: case name \'" + case + "\' not found. Skipping.")
    else:
        caseID = caseID_list[0]
        cur.execute("SELECT legal_area_id FROM legalArea WHERE name=\"" + area + "\";")
        areaID_list = cur.fetchone()
        # If the subject line is not found in the subjects table, it will be added.
        if areaID_list == None:
            # We need to make sure that we don't add a "none" entry to the table.
            if area == "None":
                return
            else:
                area_line_creation_text = "INSERT INTO legalArea (name) VALUES (\"" + area + "\");"
                execute_query(temp_connection, area_line_creation_text)
                cur.execute("SELECT legal_area_id FROM legalArea WHERE name=\"" + area + "\";")
                areaID_list = cur.fetchone()
                areaID = str(areaID_list[0])
        else:
            areaID = str(areaID_list[0])
        area_crossref_creation_text = "INSERT INTO legalAreaCrossref (caseID, leal_area_id) VALUES (" + str(caseID) + ", " + areaID + ");"
        execute_query(temp_connection, area_crossref_creation_text)

def enter_legal_area_from_xl(xl_path, data_path):
    dict = read_legal_area_from_xl(xl_path)
    for case in dict:
        area_list = dict[case]
        for area in area_list:
            insert_legal_area_crossref(data_path, area, case)

def generate_legal_area_dict(data_path):
    new_dict = dict()
    temp_connection = create_connection(data_path)
    cur = temp_connection.cursor()
    cur.execute("SELECT * FROM legalArea;")
    area_list = cur.fetchall()
    for row in area_list:
        area_id = row[0]
        list_of_areaIDs = []
        cur = temp_connection.cursor()
        cur.execute("SELECT caseID from legalAreaCrossref where leal_area_id=" + str(area_id) + ";")
        sub_case_list = cur.fetchall()
        # What gets returned are somewhat messy tuples. I've converted them to simple integers
        cleaned_up_sub_case_list = []
        for member in sub_case_list:
            value = member[0]
            cleaned_up_sub_case_list.append(value)
        new_dict[area_id] = cleaned_up_sub_case_list
    return new_dict

def read_terms_from_xl(xl_path):
    raw_list = obtain_list_from_spreadsheet(xl_path)
    case_list = sort_out_raw_entry(raw_list, "special_terms")
    for case in case_list:
        case_list[case].strip()
        terms_list = case_list[case].split(",")
        terms_list = [x for x in terms_list if len(x)>3]
        terms_list = [x.strip() for x in terms_list]
        case_list[case] = terms_list
    return case_list

def insert_terms_crossref(data_path, term, case):
    subjectID = ""
    # The first step is to take the "case" given and find that case in the "cases" table in the database, and return the "id"
    temp_connection = create_connection(data_path)
    cur = temp_connection.cursor()
    cur.execute("SELECT id FROM cases WHERE name=\"" + case + "\";")
    caseID_list = cur.fetchone()
    if caseID_list == None:
        print("Error from insert_terms_crossref: case name \'" + case + "\' not found. Skipping.")
    else:
        caseID = caseID_list[0]
        cur.execute("SELECT term_id FROM terms WHERE name=\"" + term + "\";")
        termID_list = cur.fetchone()
        # If the subject line is not found in the subjects table, it will be added.
        if termID_list == None:
            # We need to make sure that we don't add a "none" entry to the table.
            if term == "None":
                return
            else:
                term_line_creation_text = "INSERT INTO terms (name) VALUES (\"" + term + "\");"
                execute_query(temp_connection, term_line_creation_text)
                cur.execute("SELECT term_id FROM terms WHERE name=\"" + term + "\";")
                termID_list = cur.fetchone()
                termID = str(termID_list[0])
        else:
            termID = str(termID_list[0])
        term_crossref_creation_text = "INSERT INTO termsCrossref (caseID, term_id) VALUES (" + str(caseID) + ", " + termID + ");"
        execute_query(temp_connection, term_crossref_creation_text)

def enter_terms_from_xl(xl_path, data_path):
    dict = read_terms_from_xl(xl_path)
    for case in dict:
        terms_list = dict[case]
        for term in terms_list:
            insert_terms_crossref(data_path, term, case)

def find_alteration_matches(spreadsheet_path, data_path):
    whole_list = obtain_all_cases(data_path)
    whole_list_titles = ()
    for case in whole_list:
        case_name = case[1]
        whole_list_titles = whole_list_titles + (case_name,)
    wb = load_workbook(spreadsheet_path)
    ws = wb["Alterations"]
    list_of_entries = dict()
    for row in ws.iter_rows(min_row=2, max_row=1000, min_col=1, max_col=1):
        for i in row:
            if i.value == None:
                pass
            else:
                case_string = str(i.value)
                list_of_entries[case_string.strip()] = i.row
    matched_cases = dict()
    for case in whole_list_titles:
        if case in list_of_entries:
            if case in matched_cases:
                pass
            else:
                matched_cases[case] = list_of_entries[case]
        else:
            pass
    return matched_cases

def execute_alterations(xl_path, data_path):
    alter_dict = find_alteration_matches(xl_path, data_path)
    print(alter_dict)
    wb = load_workbook(xl_path)
    ws = wb["Alterations"]
    for entry in alter_dict:
        row_value = alter_dict[entry]
        row_to_examine = ws[row_value]
        case_name = entry
        for cell in row_to_examine:
            if cell.value == None or len(str(cell.value)) < 3:
                pass
            else:
                if cell.column == 2:
                    cell_contents = str(cell.value)
                    cell_contents = "\"" + cell_contents.strip() + "\""
                    print("[execute_alterations]: found change to " + case_name)
                    print("Changing name to: " + cell_contents)
                    perform_replace_alteration("cases", "name", cell_contents, case_name, data_path)
                    print("Alteration done.")
                elif cell.column == 3:
                    cell_contents = str(cell.value)
                    cell_contents.strip()
                    print("[execute_alterations]: found change to " + case_name)
                    print("Changing year to: [" + cell_contents + "]")
                    perform_replace_alteration("cases", "year", cell_contents, case_name, data_path)
                    print("Alteration done.")
                elif cell.column == 4:
                    cell_contents = str(cell.value)
                    cell_contents = "\"" + cell_contents.strip() + "\""
                    print("[execute_alterations]: found change to " + case_name)
                    print("Changing nominate citation to: " + cell_contents)
                    perform_replace_alteration("cases", "nom_cite", cell_contents, case_name, data_path)
                    print("Alteration done.")
                elif cell.column == 5:
                    cell_contents = str(cell.value)
                    cell_contents = "\"" + cell_contents.strip() + "\""
                    print("[execute_alterations]: found change to " + case_name)
                    print("Changing ER/Modern Citation to: " + cell_contents)
                    perform_replace_alteration("cases", "er_cite", cell_contents, case_name, data_path)
                    print("Alteration done.")
                elif cell.column == 6:
                    cell_contents = str(cell.value)
                    cell_contents = "\"" + cell_contents.strip() + "\""
                    print("[execute_alterations]: found change to " + case_name)
                    print("Changing court to: " + cell_contents)
                    perform_replace_alteration("cases", "court", cell_contents, case_name, data_path)
                    print("Alteration done.")
                elif cell.column == 7:
                    cell_contents = str(cell.value)
                    if "CLEAR" in cell_contents:
                        print("[execute_alterations]: found change to " + case_name)
                        print("Clearing and replacing subject alterations with: " + cell_contents)
                        perform_clear_and_replace_alteration(data_path, "subjects", case_name, cell_contents)
                    else:
                        print("[execute_alterations]: found change to " + case_name)
                        print("Adding to subjects: " + cell_contents)
                        perform_add_alteration(data_path, "subjects", case_name, cell_contents)
                        print("Alteration done.")
                elif cell.column == 8:
                    cell_contents = str(cell.value)
                    if "CLEAR" in cell_contents:
                        print("[execute_alterations]: found change to " + case_name)
                        print("Clearing and replacing authors alterations with: " + cell_contents)
                        perform_clear_and_replace_alteration(data_path, "authors", case_name, cell_contents)
                    else:
                        cell_contents = cell_contents.strip()
                        print("[execute_alterations]: found change to " + case_name)
                        print("Adding to authors: " + cell_contents)
                        perform_add_alteration(data_path, "authors", case_name, cell_contents)
                        print("Alteration done.")
                elif cell.column == 9:
                    print("Changes to cases cited")
                    print(str(row_value) + "," + str(cell.column))
                elif cell.column == 10:
                    cell_contents = str(cell.value)
                    if "CLEAR" in cell_contents:
                        print("[execute_alterations]: found change to " + case_name)
                        print("Clearing and replacing cite-in alterations with: " + cell_contents)
                        perform_clear_and_replace_alteration(data_path, "cite_in", case_name, cell_contents)
                    else:
                        cell_contents = cell_contents.strip()
                        print("[execute_alterations]: found change to " + case_name)
                        print("Adding to cited-in reference: \n" + cell_contents)
                        perform_add_alteration(data_path, "cited_in", case_name, cell_contents)
                        print("Alteration done.")
                        print(str(row_value) + "," + str(cell.column))
                elif cell.column == 11:
                    cell_contents = str(cell.value)
                    cell_contents = "\"" + cell_contents.strip() + "\""
                    print("[execute_alterations]: found change to " + case_name)
                    print("Changing Note to: \n" + cell_contents)
                    perform_replace_alteration("cases", "comment", cell_contents, case_name, data_path)
                    print("Alteration done.")
                elif cell.column == 12:
                    cell_contents = str(cell.value)
                    if "CLEAR" in cell_contents:
                        cell_contents = cell_contents.strip()
                        print("[execute_alterations]: found change to " + case_name)
                        print("Clearing and replacing legal_area with: " + cell_contents)
                        perform_clear_and_replace_alteration(data_path, "legalArea", case_name, cell_contents)
                    else:
                        cell_contents = str(cell.value)
                        cell_contents = cell_contents.strip()
                        print("[execute_alterations]: found change to " + case_name)
                        print("Adding to legal area: " + cell_contents)
                        perform_add_alteration(data_path, "legalArea", case_name, cell_contents)
                elif cell.column == 13:
                    cell_contents = str(cell.value)
                    cell_contents = "\"" + cell_contents.strip() + "\""
                    print("[execute_alterations]: found change to " + case_name)
                    print("Changing link to: \n" + cell_contents)
                    perform_replace_alteration("cases", "link", cell_contents, case_name, data_path)
                    print("Alteration done.")
                elif cell.column == 14:
                    cell_contents = str(cell.value)
                    if "CLEAR" in cell_contents:
                        perform_clear_and_replace_alteration(data_path, "terms", case_name, cell_contents)
                    else:
                        cell_contents = str(cell.value)
                        cell_contents = cell_contents.strip()
                        print("[execute_alterations]: found change to " + case_name)
                        print("Clearing and adding to terms: " + cell_contents)
                        perform_add_alteration(data_path, "terms", case_name, cell_contents)

def perform_replace_alteration(table_name, field, value, case_name, data_path):
    # First, find the ID for the case_name
    temp_connection = create_connection(data_path)
    cur = temp_connection.cursor()
    cur.execute("SELECT id FROM cases WHERE name=\"" + case_name + "\";")
    caseID = cur.fetchone()
    if caseID == None:
        return
    else:
        update_line_1 = "UPDATE " + table_name + "\n"
        update_line_2 = "SET " + field + "=" + value + "\n"
        update_line_3 = "WHERE id=" + str(caseID[0]) + ";"
        full_update_text = update_line_1 + update_line_2 + update_line_3
        current_connection = create_connection(data_path)
        execute_query(current_connection, full_update_text)

# This function is only used for the ammendment tools. It takes a string, which is the basic conents of the cell, and returns a nested list, containing [[author, comment], [author, comment]]
def read_author_comment_cell(value):
    initial_split_list = value.split(";")
    initial_split_list = [x for x in initial_split_list if len(x)>3]
    if initial_split_list[0] == None:
        return None
    else:
        total_list = []
        for citation in initial_split_list:
            citation.strip()
            author_and_comment = citation.split("[")
            author_and_comment[1] = author_and_comment[1].rstrip()
            author_and_comment[1] = author_and_comment[1].lstrip()
            author_and_comment[0] = author_and_comment[0].strip()
            author_and_comment[1] = (author_and_comment[1])[:-1]
            total_list.append(author_and_comment)
    return total_list

def add_alteration_citing_crossref(author, comment, case, data_path):
    # First, obtain the caseID.
    temp_connection = create_connection(data_path)
    cur = temp_connection.cursor()
    cur.execute("SELECT id FROM cases where name=\"" + case + "\";")
    caseID = (cur.fetchone())[0]
    # Then, check to see if there is already an entry under the "author" given.
    cur = temp_connection.cursor()
    cur.execute("SELECT authorID FROM authors where name=\"" + author + "\";")
    author_ID = (cur.fetchone())
    # If no author ID is found, a new one is added.
    if author_ID == None:
        author_line_creation_text = "INSERT INTO authors (name) VALUES (\"" + author + "\");"
        execute_query(temp_connection, author_line_creation_text)
        print("[add_alteration_citing_crossref]: I do not have the author " + author + " so I have added it.")
        author_ID = cur.execute("SELECT authorID FROM authors where name=\"" + author + "\";")
        author_ID = (cur.fetchone())
    else:
        pass
    # Then, go about the business of actually adding the crossreference.
    author_ID = author_ID[0]
    author_commment_creation_text = "INSERT INTO authorsCrossref (authorID, caseID, comment) VALUES (" + str(author_ID) + ", " + str(caseID) + ", \" " + comment + "\");"
    execute_query(temp_connection, author_commment_creation_text)



def perform_add_alteration(data_path, table_name, case_name, cell_value):
    if table_name == "subjects":
        # You first need to split up the tags
        cell_value.strip()
        subject_list = cell_value.split(" ")
        subject_list = [x for x in subject_list if len(x)>3]
        for i in subject_list:
            i.strip()
            insert_subjects_crossref(data_path, i, case_name)
    elif table_name == "legalArea":
        cell_value.strip()
        area_list = cell_value.split(" ")
        area_list = [x for x in area_list if len(x)>3]
        for i in area_list:
            i.strip()
            insert_legal_area_crossref(data_path, i, case_name)
    elif table_name == "authors":
        cell_value.strip()
        authors_list = cell_value.split(" ")
        authors_list = [x for x in authors_list if len(x)>3]
        for i in authors_list:
            i.strip()
            insert_citing_crossref(data_path, i, case_name)
    elif table_name == "terms":
        cell_value.strip()
        terms_list = cell_value.split(",")
        terms_list = [x for x in terms_list if len(x)>3]
        for i in terms_list:
            i = i.strip()
            insert_terms_crossref(data_path, i, case_name)
    elif table_name == "cited_in":
        # This whole thing was much more complicated than it needed to be, and required me to create some new functions.
        # The first step is to split the input into a list of [author, comment] lists.
        author_and_comment_list = read_author_comment_cell(cell_value)
        # Then, for each of those, another function is called up to make the relevant changes.
        for item in author_and_comment_list:
            author = item[0]
            comment = item[1]
            add_alteration_citing_crossref(author, comment, case_name, data_path)
    else:
        return


def perform_clear_and_replace_alteration(data_path, table_name, case_name, cell_value):
    if table_name == "subjects":
        temp_connection = create_connection(data_path)
        cur = temp_connection.cursor()
        cur.execute("SELECT id FROM cases where name=\"" + case_name + "\";")
        caseID = (cur.fetchone())[0]
        deletion_text = "DELETE FROM subjectsCrossref where caseID=" + str(caseID) + ";"
        execute_query(temp_connection, deletion_text)
        cell_value = cell_value.replace("CLEAR", "")
        cell_value = cell_value.strip()
        if cell_value == "":
            return
        else:
            perform_add_alteration(data_path, table_name, case_name, cell_value)
    elif table_name == "legalArea":
        temp_connection = create_connection(data_path)
        cur = temp_connection.cursor()
        cur.execute("SELECT id FROM cases where name=\"" + case_name + "\";")
        caseID = (cur.fetchone())[0]
        deletion_text = "DELETE FROM legalAreaCrossref where caseID=" + str(caseID) + ";"
        print(deletion_text)
        execute_query(temp_connection, deletion_text)
        cell_value = cell_value.replace("CLEAR", "")
        cell_value = cell_value.strip()
        if cell_value == "":
            return
        else:
            perform_add_alteration(data_path, table_name, case_name, cell_value)
    elif table_name == "authors":
        temp_connection = create_connection(data_path)
        cur = temp_connection.cursor()
        cur.execute("SELECT id FROM cases where name=\"" + case_name + "\";")
        caseID = (cur.fetchone())[0]
        deletion_text = "DELETE FROM citingCrossref where caseID=" + str(caseID) + ";"
        print(deletion_text)
        execute_query(temp_connection, deletion_text)
        cell_value = cell_value.replace("CLEAR", "")
        cell_value = cell_value.strip()
        if cell_value == "":
            return
        else:
            perform_add_alteration(data_path, table_name, case_name, cell_value)
    elif table_name == "terms":
        temp_connection = create_connection(data_path)
        cur = temp_connection.cursor()
        cur.execute("SELECT id FROM cases where name=\"" + case_name + "\";")
        caseID = (cur.fetchone())[0]
        deletion_text = "DELETE FROM termsCrossref where caseID=" + str(caseID) + ";"
        print(deletion_text)
        execute_query(temp_connection, deletion_text)
        cell_value = cell_value.replace("CLEAR", "")
        cell_value = cell_value.strip()
        if cell_value == "":
            return
        else:
            perform_add_alteration(data_path, table_name, case_name, cell_value)
    elif table_name == "cited_in":
        temp_connection = create_connection(data_path)
        cur = temp_connection.cursor()
        cur.execute("SELECT id FROM cases where name=\"" + case_name + "\";")
        caseID = (cur.fetchone())[0]
        deletion_text = "DELETE FROM authorsCrossref where caseID=" + str(caseID) + ";"
        print(deletion_text)
        execute_query(temp_connection, deletion_text)
        cell_value = cell_value.replace("CLEAR", "")
        cell_value = cell_value.strip()
        if cell_value == "":
            return
        else:
            perform_add_alteration(data_path, table_name, case_name, cell_value)
    else:
        return


def produce_listed_field_display_string(data_path, id, entry_type):
    # This function produces a list which goes into a spreadsheet to show anything that appears as a "list" from the spreadsheet.
    if entry_type == "subject":
        temp_connection = create_connection(data_path)
        cur = temp_connection.cursor()
        cur.execute("SELECT subjectsID FROM subjectsCrossref WHERE caseID=" + str(id) + ";")
        subject_ids = cur.fetchall()
        list = ""
        n = 0
        for id in subject_ids:
            if n < len(subject_ids)-1:
                list = list + str(id[0]) + ", "
            else:
                list = list + str(id[0])
            n = n + 1
        cur.execute("SELECT name from subjects WHERE id IN(" + list + ");")
        subject_names = cur.fetchall()
        return_list = ""
        for tag in subject_names:
            return_list = return_list + tag[0] + " "
    elif entry_type == "legal_area":
        temp_connection = create_connection(data_path)
        cur = temp_connection.cursor()
        cur.execute("SELECT leal_area_id FROM legalAreaCrossref WHERE caseID=" + str(id) + ";")
        area_ids = cur.fetchall()
        list = ""
        n = 0
        for id in area_ids:
            if n < len(area_ids)-1:
                list = list + str(id[0]) + ", "
            else:
                list = list + str(id[0])
            n = n + 1
        cur.execute("SELECT name from legalArea WHERE legal_area_id IN(" + list + ");")
        area_names = cur.fetchall()
        return_list = ""
        for tag in area_names:
            return_list = return_list + tag[0] + " "
    elif entry_type == "terms":
        temp_connection = create_connection(data_path)
        cur = temp_connection.cursor()
        cur.execute("SELECT term_id FROM termsCrossref WHERE caseID=" + str(id) + ";")
        term_ids = cur.fetchall()
        list = ""
        n = 0
        for id in term_ids:
            if n < len(term_ids)-1:
                list = list + str(id[0]) + ", "
            else:
                list = list + str(id[0])
            n = n + 1
        cur.execute("SELECT name from terms WHERE term_id IN(" + list + ");")
        term_names = cur.fetchall()
        return_list = ""
        for tag in term_names:
            return_list = return_list + tag[0] + ", "
    elif entry_type == "authors":
        temp_connection = create_connection(data_path)
        cur = temp_connection.cursor()
        cur.execute("SELECT DISTINCT author_cited FROM citingCrossref WHERE caseID=" + str(id) + ";")
        authors_ids = cur.fetchall()
        list = ""
        n = 0
        for id in authors_ids:
            if n < len(authors_ids)-1:
                list = list + str(id[0]) + ", "
            else:
                list = list + str(id[0])
            n = n + 1
        cur.execute("SELECT name from citing WHERE author_citing_id IN(" + list + ");")
        author_names = cur.fetchall()
        return_list = ""
        for tag in author_names:
            return_list = return_list + tag[0] + " "
    elif entry_type == "cited_in":
        temp_connection = create_connection(data_path)
        cur = temp_connection.cursor()
        cur.execute("SELECT DISTINCT authorID, comment FROM authorsCrossref WHERE caseID=" + str(id) + ";")
        cited_ids_and_comments = cur.fetchall()
        return_list = ""
        for pair in cited_ids_and_comments:
            cur.execute("SELECT name FROM authors WHERE authorID=" + str(pair[0]) + ";")
            author = cur.fetchone()[0]
            comment = pair[1]
            string_to_add = author.strip() + "[" + comment.strip() + "]; "
            return_list = return_list + string_to_add
    else:
        return
    return return_list


def enter_all_cases_into_sheet(spreadsheet_path, data_path):
    header = ["ID", "Name", "Year", "Nom Cite", "ER Cite", "Court", "Link", "Comment", "Subjects", "Legal Area", "Authors Cited", "Special Terms", "Cited In"]
    list_of_cases = obtain_all_cases(data_path)
    wb = load_workbook(spreadsheet_path)
    page = wb["Existing_Entries"]
    page.delete_rows(1, page.max_row)
    page.append(header)
    for case in list_of_cases:
        subjects = produce_listed_field_display_string(data_path, case[0], "subject")
        case += (subjects,)
        areas = produce_listed_field_display_string(data_path, case[0], "legal_area")
        case += (areas,)
        authors = produce_listed_field_display_string(data_path, case[0], "authors")
        case += (authors,)
        terms = produce_listed_field_display_string(data_path, case[0], "terms")
        case += (terms,)
        cited_in = produce_listed_field_display_string(data_path, case[0], "cited_in")
        case += (cited_in,)
        page.append(case)
    wb.save(filename=spreadsheet_path)
    #print(list_of_cases)


def quick_start(data_name, xl_name):
    make_new_table(data_name, create_case_table)
    make_new_table(data_name, create_subject_table)
    make_new_table(data_name, create_subject_crosref_table)
    make_new_table(data_name, create_authors_table)
    make_new_table(data_name, create_authors_crossref_table)
    make_new_table(data_name, create_citing_table)
    make_new_table(data_name, create_citing_table_crossref)
    make_new_table(data_name, create_legal_area_table)
    make_new_table(data_name, create_legal_area_table_crossref)
    make_new_table(data_name, create_terms_table)
    make_new_table(data_name, create_terms_table_crossref)
    temp_connection = create_connection(data_name)
    insert_cases_from_xl(xl_name, data_name)
    enter_subject_tags_from_xl(xl_name, data_name)
    enter_author_comments_from_xl(xl_name, data_name)
    enter_citing_from_xl(xl_name, data_name)
    enter_legal_area_from_xl(xl_name, data_name)
    enter_terms_from_xl(xl_name, data_name)

#quick_start("data2.db", "Test_Input.xlsx")

SPREADSHEET = str(sys.argv[1])

DATABASE = str(sys.argv[2])

MODE = int(sys.argv[3])

print(SPREADSHEET)

print(DATABASE)

print(MODE)

if MODE == 0:
    print("Entering data into existing database.")
    temp_connection = create_connection(DATABASE)
    insert_cases_from_xl(SPREADSHEET, DATABASE)
    enter_subject_tags_from_xl(SPREADSHEET, DATABASE)
    enter_author_comments_from_xl(SPREADSHEET, DATABASE)
    enter_citing_from_xl(SPREADSHEET, DATABASE)
    enter_legal_area_from_xl(SPREADSHEET, DATABASE)
    enter_terms_from_xl(SPREADSHEET, DATABASE)
    execute_alterations(SPREADSHEET, DATABASE)
    enter_all_cases_into_sheet(SPREADSHEET, DATABASE)
elif MODE == 1:
    print("Starting a new database.")
    quick_start(DATABASE, SPREADSHEET)
else:
    pass
